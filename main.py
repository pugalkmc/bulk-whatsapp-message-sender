# import pywhatkit
# import time
# import openpyxl
# import random

# def send_whatsapp_message(phone_number, message_text, image_path):
#     try:
#         # print("Please log in to WhatsApp Web and keep it open.")
#         time.sleep(5)
#         pywhatkit.sendwhats_image(phone_number, image_path, message_text)
#         print(f"Message and image sent to {phone_number}.")
#     except Exception as e:
#         print(f"Error sending message and image to {phone_number}: {e}")

# def read_numbers_from_excel(excel_file_path):
#     try:
#         workbook = openpyxl.load_workbook(excel_file_path)
#         worksheet = workbook.active
#         phone_numbers = []
#         for row in worksheet.iter_rows(min_row=1):
#             phone_number = row[0].value
#             phone_numbers.append(str(int(phone_number)))
#         return phone_numbers
#     except Exception as e:
#         print(f"Error reading phone numbers from Excel file: {e}")
#         return []

# # Main program
# if __name__ == "__main__":
#     excel_file_path = "original.xlsx"
#     message_text = "Final testing"
#     image_path = "image.jpg"

#     phone_numbers = read_numbers_from_excel(excel_file_path)
#     print(phone_numbers)
#     # phone_numbers = ["9344776097", "7639262994"]
#     if phone_numbers:
#         for phone_number in phone_numbers:
#             if len(phone_number) >= 10 and phone_number.isdigit():
#                 delay = random.randint(5, 12)
#                 time.sleep(delay)
#                 send_whatsapp_message("+91" + phone_number, message_text, image_path)
#             else:
#                 print(f"Skipping invalid phone number: {phone_number}")
#     else:
#         print("No phone numbers found in the Excel file.")
import pywhatkit
import time
import openpyxl
import random

def send_whatsapp_message(phone_number, message_text, image_path):
    try:
        # print("Please log in to WhatsApp Web and keep it open.")
        time.sleep(5)
        pywhatkit.sendwhats_image(phone_number, image_path, message_text)
        print(f"Message and image sent to {phone_number}.")
        return True
    except Exception as e:
        print(f"Error sending message and image to {phone_number}: {e}")
        return False

def read_numbers_from_excel(excel_file_path):
    try:
        workbook = openpyxl.load_workbook(excel_file_path)
        worksheet = workbook.active
        phone_numbers = []
        for row in worksheet.iter_rows(min_row=1):
            phone_number = row[0].value
            phone_numbers.append(str(int(phone_number)))
        return phone_numbers
    except Exception as e:
        print(f"Error reading phone numbers from Excel file: {e}")
        return []

def write_to_excel(file_path, data):
    try:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for item in data:
            worksheet.append([item])
        workbook.save(file_path)
        print(f"Data written to {file_path}.")
    except Exception as e:
        print(f"Error writing data to Excel file: {e}")

# Main program
if __name__ == "__main__":
    excel_file_path = "duplicate.xlsx"
    message_text = "Final testing"
    image_path = "image.jpg"

    success_numbers = []
    failed_numbers = []
    invalid_numbers = []

    phone_numbers = read_numbers_from_excel(excel_file_path)
    print(phone_numbers)

    if phone_numbers:
        for phone_number in phone_numbers:
            if len(phone_number) >= 10 and phone_number.isdigit():
                delay = random.randint(5, 12)
                time.sleep(delay)
                success = send_whatsapp_message("+91" + phone_number, message_text, image_path)
                if success:
                    success_numbers.append(phone_number)
                else:
                    failed_numbers.append(phone_number)
            else:
                print(f"Skipping invalid phone number: {phone_number}")
                invalid_numbers.append(phone_number)

        # Write successful, failed, and invalid numbers to separate Excel files
        write_to_excel("success.xlsx", success_numbers)
        write_to_excel("failed.xlsx", failed_numbers)
        write_to_excel("invalid.xlsx", invalid_numbers)

        # Retry sending messages for failed numbers
        retry_failed_numbers = read_numbers_from_excel("failed.xlsx")
        for phone_number in retry_failed_numbers:
            success = send_whatsapp_message("+91" + phone_number, message_text, image_path)
            if success:
                success_numbers.append(phone_number)

        # Append successfully sent numbers to success Excel file
        write_to_excel("success.xlsx", success_numbers)
